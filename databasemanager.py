import pandas as pd
import mysql.connector
import os
from dotenv import load_dotenv
from databaseconnection import DatabaseConnection
from queryexecutor import QueryExecutor
import openpyxl


class DatabaseManager:
    def __init__(self, db_config):
        self.db_connection = DatabaseConnection(db_config)
        self.query_executor = QueryExecutor(self.db_connection)

    def connect(self):
        self.db_connection.connect()

    def disconnect(self):
        self.db_connection.close()

    # ******************* FETCH FUNCTIONS *******************
    def fetch_name_by_m13(self, m13id):
        query = """SELECT displayname
        FROM smarter.fdppops
        WHERE m13id = %s;
        """
        return self.query_executor.execute_query(query, params=(m13id,))

    def fetch_name_by_ticketid(self, ticket_id):
        query = """SELECT displayname
        FROM smarter.fdppops
        WHERE ticketid = %s;
        """
        return self.query_executor.execute_query(query, params=(ticket_id,))

    def fetch_messages_by_ticketid(self, ticket_id):
        query = """ SELECT TicketID, DateReceivedUTC, BodyHTML, MessageDirection
        FROM smarter.st_ticketmessages 
        WHERE TicketID = %s;
        """
        return self.query_executor.execute_query(query, params=(ticket_id,))

    # Fetch a list of IDs based on the parameters (LIMIT amount of records from the year YEAR)
    def fetch_id_list(self, limit, year):
        query = """SELECT *
        FROM smarter.fdppops
        WHERE mediastart BETWEEN %(start_date)s AND %(end_date)s
        ORDER BY ticketid
        LIMIT %(limit)s;
        """
        params = {
            "start_date": f"{year}-01-01",
            "end_date": f"{year}-12-31",
            "limit": limit,
        }
        return self.query_executor.execute_query(query, params=params)

    # ******************* OUTPUT FUNCTIONS *******************
    def save_conversation_as_txt(self, df, ticket_id, contact_name):
        query = """SELECT * FROM smarter.fdppops WHERE TicketID = %s;
        """
        temp = self.query_executor.execute_query(query, params=(ticket_id,))
        m13id = temp["m13id"].to_string(index=False, header=False)
        folder_path = f"messages"  # NOTE: default folder name
        if not os.path.exists(folder_path):
            os.makedirs(folder_path)
        file_name = f"{folder_path}/{m13id}.txt"

        messages = []
        for _, row in df.iterrows():
            if row["MessageDirection"] == 0:
                msg = f"{contact_name}: " + str(row["BodyHTML"])
            elif row["MessageDirection"] == 1:
                msg = "AGENT: " + str(row["BodyHTML"])
            else:
                continue  # this should never happen
            messages.append(msg)

        conversation = "\n".join(messages)
        with open(file_name, "w") as file:
            file.write(conversation)

    def save_control_data_to_excel(self, df, year):
        control_df = df[
            [
                "m13id",
                "displayname",
                "agegender",
                "attitude",
                "level",
                "kotakab",
                "district",
                "statushp",
            ]
        ]
        control_df[["Age (now)", "Gender"]] = control_df["agegender"].str.split(
            "/", expand=True
        )
        control_df = control_df.drop(columns=["agegender"])

        print(control_df)
        control_df.rename(
            columns={
                "m13id": "M13 ID",
                "displayname": "Name",
                "attitude": "Attitude",
                "level": "Level",
                "kotakab": "City",
                "district": "Kecamatan",
                "statushp": "Status HP",
            },
            inplace=True,
        )

        # save to excel
        excel_file_name = "control1.xlsx"

        # Define headers
        headers = [
            "M13 ID",
            "Name",
            "Attitude",
            "Level",
            "City",
            "Kecamatan",
            "Status HP",
            "Age (now)",
            "Gender",
        ]

        if os.path.exists(excel_file_name):
            # Load the existing workbook and select the active worksheet
            wb = openpyxl.load_workbook(excel_file_name)
            ws = wb.active

            # Ensure the headers are consistent
            existing_headers = [cell.value for cell in ws[1]]
            if existing_headers != headers:
                raise ValueError("Existing headers do not match the required headers!")
        else:
            # Create a new workbook and set up the sheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Control Data"
            ws.append(headers)  # Add headers to the new sheet

        # Append new data
        for row in control_df.itertuples(index=False, name=None):
            ws.append(row)

        # Save the workbook
        wb.save(excel_file_name)


# Example usage
if __name__ == "__main__":
    load_dotenv()
    db_config = {
        "host": os.getenv("DB_HOST"),
        "user": os.getenv("DB_USER"),
        "password": os.getenv("DB_PASSWORD"),
        "database": os.getenv("DB_NAME"),
        "port": os.getenv("DB_PORT"),
    }

    db_manager = DatabaseManager(db_config)
    db_manager.connect()

    try:
        # NOTE: Modify the year here
        year = 2024
        limit = 500

        # Fetch 100 data from the year 2024
        df = db_manager.fetch_id_list(limit=limit, year=year)

        print(df.columns)  # debug

        # Save the "control data" to an excel sheet
        db_manager.save_control_data_to_excel(df=df, year=year)

        # Save the conversations as plain texts
        # id_list = df["ticketid"].tolist()
        # for id_value in id_list:
        #     df = db_manager.fetch_messages_by_ticketid(id_value)
        #     if not df.empty:
        #         name = db_manager.fetch_name_by_ticketid(id_value).to_string(
        #             index=False, header=False
        #         )
        #         db_manager.save_conversation_as_txt(
        #             df=df, ticket_id=id_value, contact_name=name
        #         )
        #     else:
        #         print(f"No records found for ID: {id_value}")
    finally:
        db_manager.close()
