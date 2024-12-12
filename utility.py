from contact import Contact
import json
import logging
import os
import openpyxl
import re
from bs4 import BeautifulSoup
from typing import List
from rapidfuzz import process


def setup_logging(logfile: str):
    if not logfile.endswith(".log"):
        logfile = f"{logfile}.log"
    logging.basicConfig(
        level=logging.DEBUG,  # Set the logging level to DEBUG to capture all messages
        format="%(asctime)s - %(levelname)s - %(message)s",
        handlers=[
            logging.FileHandler(logfile),  # Log file name
            logging.StreamHandler(),  # Continue to print to console as well
        ],
    )
    logging.getLogger("httpcore").setLevel(logging.WARNING)
    logging.getLogger("httpx").setLevel(logging.WARNING)


def parse_json_to_contact(json_data):
    try:
        # json_data = validate_and_fix_json(json_data=json_data)
        data = json.loads(json_data)

        contact_info = {
            "name": data.get("name_result", ""),
            "phone_number": data.get("handphone_result", ""),
            "gender": data.get("gender_result", ""),
            "age": data.get("age_result", ""),
            "education": data.get("education_result", ""),
            "occupation": data.get("occupation_result", ""),
            "marriage": data.get("marriage_result", ""),
            "attitude": data.get("attitude_result", ""),
            "persona": data.get("persona_initial_theme", ""),
            "summary": data.get("comments_idn", ""),
            "extra_info": data.get("extra_info", ""),
            "status_hp": data.get("status_hp_result", ""),
            "suku": data.get("suku_result", ""),
            "province": data.get("address_province_result", ""),
            "city": data.get("address_city_result", ""),
            "kecamatan": data.get("address_kecamatan_result", ""),
            "address": data.get("address_detail_result", ""),
        }
        return Contact(**contact_info)
    except json.decoder.JSONDecodeError as e:
        # Extract the line and column that caused the error
        lines = json_data.splitlines()
        error_line = lines[
            e.lineno - 1
        ]  # Get the line where the error occurred (e.lineno is 1-based)
        column_error = e.colno  # Column number where the error occurred

        # Log the error with detailed information, including the exact line and column content
        logging.error(
            "Failed to parse JSON data into a Contact object. "
            "The provided data might be malformed or incomplete. "
            f"Error: {e.msg} at line {e.lineno}, column {e.colno}. "
            f"Problematic line: '{error_line.strip()}'"
            f"\n{' ' * (column_error - 1)}^ <-- Error at column {column_error}. "
            "Input data (up to 200 chars): %s",
            json_data[
                :200
            ],  # Log the first 200 characters of the JSON data for context
        )
    except KeyError as e:
        # If a key is missing from the expected JSON structure, log that error explicitly
        logging.error(
            f"Missing expected key in the parsed JSON data: {e}. "
            f"Input data: {json_data[:200]}"  # Log the first 200 characters of the JSON data
        )
    except Exception as e:
        # Catch any other general errors and log them
        logging.error(
            f"An unexpected error occurred when processing the JSON data. Error: {e}. "
            f"Input data: {json_data[:200]}"  # Log the first 200 characters of the JSON data
        )
    return None  # This will cause error


EXCEL_HEADERS = [
    "M13 ID",
    "Name",
    "Attitude",
    "Handphone",
    "Persona",
    "Status HP",
    "Suku",
    "Gender",
    "Province",
    "Age (now)",
    "Level",
    "Education",
    "City",
    "Occupation",
    "Kecamatan",
    "Marriage",
    "Address",
    "Extra Info",
    "Comments",
]


def contacts_to_excel(contacts, excel_file_name):
    if os.path.exists(excel_file_name):
        wb = openpyxl.load_workbook(excel_file_name)
        ws = wb.active
        headers = [cell.value for cell in ws[1]]  # Read the first row as headers
    else:
        # Create a new Excel workbook and select the active worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Contacts"
        headers = EXCEL_HEADERS
        ws.append(headers)

    for contact in contacts:
        if contact is None:
            logging.error("Contact is None in contacts_to_excel. Skipping...")
            continue
        row_data = [
            str(contact.id),
            str(contact.name),
            str(contact.attitude),
            str(contact.phone_number),
            str(contact.persona),
            str(contact.status_hp),
            str(contact.suku),
            str(contact.gender),
            str(contact.province),
            str(contact.age),
            str(contact.level),  # 'level' is now directly accessed as a property
            str(contact.education),
            str(contact.city),
            str(contact.occupation),
            str(contact.kecamatan),
            str(contact.marriage),
            str(contact.address),
            str(contact.extra_info),
            str(contact.summary),
        ]
        ws.append(row_data)

    wb.save(excel_file_name)


# This function extracts and returns the ID in the format "<A-Z> <4 digit numbers>" from a file path
def get_file_id(file_path):
    pattern = r"([A-Z])\s(\d{4})"
    file_name = os.path.basename(file_path)
    logging.debug(f"Extracted file name: {file_name}")
    match = re.search(pattern, file_name)
    if match:
        logging.info(f"Match found: {match.group(0)}")
        return f"{match.group(1)} {match.group(2)}"
    else:
        logging.info("No match found.")
        return None


def clean_json(input_string):
    """
    Removes everything before the first `{` and after the last `}` in the given string.

    :param input_string: The string to clean.
    :return: The cleaned string or an empty string if `{` or `}` is missing.
    """
    try:
        start = input_string.index("{")
        end = input_string.rindex("}")

        cleaned_string = input_string[start : end + 1]

        # Check for a comma before the closing `}`
        # Strip spaces from the end to find the character before the `}`
        trimmed_before_closing = cleaned_string[: end - start].rstrip()

        # If the character before the final `}` is a comma, remove it
        if trimmed_before_closing[-1] == ",":
            # Remove the trailing comma and restore the `}`
            cleaned_string = trimmed_before_closing[:-1] + "}"

        return cleaned_string

    except ValueError:
        # If `{` or `}` is not found, return an empty string
        return ""


# This function extracts and returns the ID in the format "<A-Z> <4 digit numbers>" from a file path
def get_file_id(file_path):
    pattern = r"([A-Z])\s(\d{4})"
    file_name = os.path.basename(file_path)
    logging.debug(f"Extracted file name: {file_name}")
    match = re.search(pattern, file_name)
    if match:
        logging.info(f"Match found: {match.group(0)}")
        return f"{match.group(1)} {match.group(2)}"
    else:
        logging.info("No match found.")
        return None


def validate_excel(filename: str):
    if ".xlsx" not in filename:
        return filename + ".xlsx"
    else:
        return filename


def clean_html_styling(conversation_text):
    soup = BeautifulSoup(conversation_text, "html.parser")
    return soup.get_text()


def fuzzy_search(keyword: str, column: List[str]):
    match = process.extractOne(keyword, column)
    if match:
        return match[0]
    return ""
